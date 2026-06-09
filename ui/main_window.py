import customtkinter as ctk
from tkinter import filedialog, messagebox
import pandas as pd
import os
import platform
import subprocess
import re
import queue
import threading

import config
from config import LITE_MODE
from models.id_parser import (
    parse_id_card, safe_filename,
    extract_placeholders_from_template,
)
from models.fee_calculator import get_calculator
from models.formula_engine import validate_formula
from services.excel_processor import process_raw_excel
from services.word_generator import generate_documents
from ui.widgets import TwoWayScrollableFrame

# ====================== 主题设置 ======================
ctk.set_appearance_mode("light")
ctk.set_default_color_theme("blue")

class PaymentApp(ctk.CTk):
    def __init__(self):
        super().__init__()
        self.preview_inner = None
        self.title(config.WINDOW_TITLE)
        self.geometry(config.WINDOW_SIZE)
        self.minsize(1000, 700)
        self.resizable(True, True)
        self.cancel_requested = False
        self._dark_mode = False

        self.placeholders = []
        self.word_path = ""
        self.original_excel_path = ""
        self.processed_excel_path = ""
        self.df = None
        self.check_vars = []
        self.output_folder = ""

        self.formulas = []  # 自定义公式列表
        self.auto_generated_columns = config.AUTO_GENERATED_COLUMNS.copy()
        self.template_type = ctk.StringVar(value=config.TEMPLATE_TYPES[0])
        self.preview_extra_columns = config.DEFAULT_PREVIEW_COLUMNS.copy()

        # 侧边栏 + 主区域布局
        self.sidebar = ctk.CTkFrame(self, width=config.SIDEBAR_WIDTH, corner_radius=0)
        self.sidebar.pack(side="left", fill="y")
        self.sidebar.pack_propagate(False)

        self.main_area = ctk.CTkFrame(self, fg_color="transparent")
        self.main_area.pack(side="left", fill="both", expand=True)
        self.main_area.grid_columnconfigure(0, weight=1)
        for i in range(8):
            self.main_area.grid_rowconfigure(i, weight=0)
        self.main_area.grid_rowconfigure(6, weight=1)

        self._build_sidebar()
        self._build_main()
        self._sync_theme_label()

        self.progress_queue = queue.Queue()
        self.after(100, self._process_queue)

    # ====================== 配色辅助 ======================

    @property
    def colors(self):
        return config.DARK if self._dark_mode else config.LIGHT

    def toggle_theme(self):
        """切换深色/浅色主题"""
        self._dark_mode = not self._dark_mode
        new_mode = "dark" if self._dark_mode else "light"
        ctk.set_appearance_mode(new_mode)
        self._refresh_theme_colors()

    def _refresh_theme_colors(self):
        """刷新所有控件的主题颜色（侧边栏始终深色，不刷新）"""
        c = self.colors
        self.main_area.configure(fg_color=c['main_bg'])
        # 刷新卡片
        for card in [self._file_card, self._template_card, self._action_card,
                     self._ctrl_card, self._preview_card]:
            card.configure(fg_color=c['card_bg'], border_color=c['card_border'])
        # 刷新输入框
        self.word_entry.configure(border_color=c['entry_border'], fg_color=c['entry_bg'])
        self.excel_entry.configure(border_color=c['entry_border'], fg_color=c['entry_bg'])
        self.processed_entry.configure(border_color=c['entry_border'], fg_color=c['entry_bg'])
        self.folder_entry.configure(border_color=c['entry_border'], fg_color=c['entry_bg'])
        # 刷新自定义滚动框（Canvas + 内层 Frame 统一底色）
        if hasattr(self, 'preview_scroll') and self.preview_scroll:
            self.preview_scroll.canvas.configure(bg=c['card_bg'])
            self.preview_scroll.canvas_bg = c['card_bg']
            self.preview_scroll.inner_frame.configure(fg_color=c['card_bg'])
        # 刷新状态标签
        self.status_label.configure(text_color=c['text_secondary'])
        # 重新加载预览（如果已加载数据）
        if self.df is not None and not self.df.empty:
            self.load_preview()

    # ====================== 侧边栏 ======================

    def _build_sidebar(self):
        c = self.colors
        # 侧边栏始终是深色底，用 DARK 色系保证文字可见
        sd = config.DARK
        self.sidebar.configure(fg_color=sd['sidebar_bg'])

        # Logo 区域
        logo_frame = ctk.CTkFrame(self.sidebar, fg_color="transparent")
        logo_frame.pack(fill="x", padx=18, pady=(26, 14))

        ctk.CTkLabel(
            logo_frame, text="智汇文枢",
            font=ctk.CTkFont(family="Microsoft YaHei", size=20, weight="bold"),
            text_color="#ffffff",
        ).pack(anchor="w")

        ctk.CTkLabel(
            logo_frame, text="文书自动填充工具",
            font=ctk.CTkFont(size=10),
            text_color=sd['sidebar_text'],
        ).pack(anchor="w", padx=(1, 0), pady=(2, 0))

        # 分隔线
        ctk.CTkFrame(self.sidebar, height=1, fg_color=sd['card_border']).pack(fill="x", padx=16, pady=(0, 10))

        # 导航按钮（带左侧指示条）
        nav_btns = [
            ("使用教程", self.show_tutorial),
            ("关于软件", self.show_about),
        ]
        self._nav_indicators = []
        for text, cmd in nav_btns:
            row = ctk.CTkFrame(self.sidebar, fg_color="transparent", height=38)
            row.pack(fill="x", padx=6, pady=1)
            row.pack_propagate(False)

            indicator = ctk.CTkFrame(row, width=3, fg_color="transparent", corner_radius=2)
            indicator.pack(side="left", fill="y", padx=(4, 0), pady=6)
            self._nav_indicators.append(indicator)

            btn = ctk.CTkButton(
                row, text=text,
                fg_color="transparent",
                text_color=sd['sidebar_text'],
                hover_color=sd['sidebar_hover'],
                anchor="w", height=36,
                font=ctk.CTkFont(size=15),
                corner_radius=6,
                command=cmd,
            )
            btn.pack(side="left", fill="both", expand=True)

            # hover: 指示条亮起 + 文字变白
            def make_hover_enter(ind=indicator, b=btn):
                def _enter(_e):
                    ind.configure(fg_color=sd['sidebar_active'])
                    b.configure(text_color="#ffffff")
                return _enter
            def make_hover_leave(ind=indicator, b=btn):
                def _leave(_e):
                    ind.configure(fg_color="transparent")
                    b.configure(text_color=sd['sidebar_text'])
                return _leave
            btn.bind("<Enter>", make_hover_enter(), add="+")
            btn.bind("<Leave>", make_hover_leave(), add="+")
            row.bind("<Enter>", make_hover_enter(), add="+")
            row.bind("<Leave>", make_hover_leave(), add="+")

        # 弹簧
        ctk.CTkFrame(self.sidebar, fg_color="transparent").pack(fill="both", expand=True)

        # 底部主题切换
        bottom_frame = ctk.CTkFrame(self.sidebar, fg_color=sd['sidebar_hover'], corner_radius=8)
        bottom_frame.pack(fill="x", padx=10, pady=12)

        self.theme_label = ctk.CTkLabel(
            bottom_frame, text="深色模式",
            font=ctk.CTkFont(size=13, weight="bold"),
            text_color=sd['sidebar_text'],
        )
        self.theme_label.pack(side="left", padx=(12, 0), pady=10)

        self.theme_switch = ctk.CTkSwitch(
            bottom_frame, text="",
            command=self._on_theme_switch,
            progress_color=sd['accent'],
            button_color="#ffffff",
            button_hover_color=sd['sidebar_active'],
            width=40,
        )
        self.theme_switch.pack(side="right", padx=(0, 10), pady=10)

    def _on_theme_switch(self):
        self.toggle_theme()
        self.theme_label.configure(
            text="深色模式" if self._dark_mode else "浅色模式"
        )

    # 初始化时同步标签（默认浅色）
    def _sync_theme_label(self):
        self.theme_label.configure(
            text="深色模式" if self._dark_mode else "浅色模式"
        )

    # ====================== 消息队列处理 ======================

    def _process_queue(self):
        try:
            while True:
                msg = self.progress_queue.get_nowait()
                if msg['type'] == 'progress':
                    self.progressbar.set(msg['value'])
                    self.progress_label.configure(text=msg['text'])
                elif msg['type'] == 'error':
                    messagebox.showerror("错误", msg['text'])
                elif msg['type'] == 'warning':
                    messagebox.showwarning("警告", msg['text'])
                elif msg['type'] == 'done':
                    if 'df' in msg:
                        self.df = msg['df']
                        self.after(0, self.load_preview)
                        messagebox.showinfo("处理完成", msg['text'])
                        self.status_label.configure(text=msg.get('status', '处理完成'))
                        # 所有模板：显示变量验证
                        self.after(500, self._run_variable_validation)
                    else:
                        messagebox.showinfo("完成", msg['text'])
                        count = msg.get('count', 0)
                        if count > 0:
                            self.progress_queue.put({'type': 'open_folder', 'path': msg.get('output_dir', '')})
                    if hasattr(self, 'progress_frame'):
                        self.progress_frame.grid_remove()
                    self.process_btn.configure(state="normal")
                    self.generate_btn.configure(state="normal")
                    self.preview_btn.configure(state="normal")
                    if hasattr(self, 'cancel_btn'):
                        self.cancel_btn.configure(state="disabled")
                elif msg['type'] == 'enable_buttons':
                    self.process_btn.configure(state="normal")
                    self.generate_btn.configure(state="normal")
                    self.preview_btn.configure(state="normal")
                    if hasattr(self, 'progress_frame'):
                        self.progress_frame.grid_remove()
                    if hasattr(self, 'cancel_btn'):
                        self.cancel_btn.configure(state="disabled")
                elif msg['type'] == 'cancelled':
                    messagebox.showinfo("已取消", msg['text'])
                    self.status_label.configure(text="操作已取消")
                    self.process_btn.configure(state="normal")
                    self.generate_btn.configure(state="normal")
                    self.preview_btn.configure(state="normal")
                    if hasattr(self, 'progress_frame'):
                        self.progress_frame.grid_remove()
                    if hasattr(self, 'cancel_btn'):
                        self.cancel_btn.configure(state="disabled")
                    self.cancel_requested = False
                elif msg['type'] == 'open_folder':
                    self.open_folder(msg['path'])
        except queue.Empty:
            pass
        finally:
            self.after(100, self._process_queue)

    def _process_excel_worker(self, save_path, id_config=None):
        try:
            placeholders = self.placeholders if hasattr(self, 'placeholders') else []

            def progress_cb(value, text):
                self.progress_queue.put({'type': 'progress', 'value': value, 'text': text})

            def cancel_cb():
                return self.cancel_requested

            df = process_raw_excel(
                input_path=self.original_excel_path,
                output_path=save_path,
                template_type=self.template_type.get(),
                formulas=self.formulas,
                placeholders=placeholders,
                progress_callback=progress_cb,
                cancel_checker=cancel_cb,
                id_config=id_config,
            )

            self.progress_queue.put({
                'type': 'done',
                'text': f"新 Excel 已生成：\n{save_path}",
                'status': f"处理完成！已保存至：\n{save_path}",
                'df': df,
                'template_type': self.template_type.get(),
            })

        except InterruptedError:
            self.progress_queue.put({'type': 'cancelled', 'text': "处理已取消"})
            return
        except Exception as e:
            self.progress_queue.put({'type': 'error', 'text': f"处理出错：\n{str(e)}\n\n请检查原始 Excel 格式或列名是否正确。"})
        finally:
            if not self.cancel_requested:
                self.progress_queue.put({'type': 'enable_buttons'})

    def cancel_task(self):
        self.cancel_requested = True
        self.cancel_btn.configure(state="disabled")
        self.progress_label.configure(text="正在取消...")

    def choose_preview_columns(self):
        fixed = ['序号', '业主姓名']
        available = [col for col in self.df.columns if col not in fixed and col not in self.auto_generated_columns]
        if not available:
            messagebox.showinfo("提示", "没有可预览的额外列，仅显示序号和业主姓名")
            return []

        c = self.colors
        dialog = ctk.CTkToplevel(self)
        dialog.title("选择预览列")
        dialog.geometry("400x500")
        dialog.transient(self)
        dialog.grab_set()
        dialog.configure(fg_color=c['main_bg'])

        check_vars = {}
        for col in available:
            var = ctk.BooleanVar(value=(col in self.preview_extra_columns))
            check_vars[col] = var

        scroll = ctk.CTkScrollableFrame(dialog, label_text="请勾选要额外显示的列（至少一个）",
                                         fg_color=c['card_bg'])
        scroll.pack(fill="both", expand=True, padx=20, pady=20)

        for col in available:
            ctk.CTkCheckBox(scroll, text=col, variable=check_vars[col],
                            text_color=c['text_primary']).pack(anchor="w", pady=2)

        max_cols = config.MAX_PREVIEW_EXTRA_COLUMNS

        def on_ok():
            selected = [col for col, var in check_vars.items() if var.get()]
            if len(selected) > max_cols:
                messagebox.showwarning("提示", f"最多只能选择 {max_cols} 列，请重新选择！")
                return
            self.preview_extra_columns = selected
            dialog.destroy()
            self.load_preview()

        def on_cancel():
            dialog.destroy()

        btn_frame = ctk.CTkFrame(dialog, fg_color="transparent")
        btn_frame.pack(fill="x", padx=20, pady=10)
        ctk.CTkButton(btn_frame, text="确定", fg_color=c['accent'],
                      hover_color=c['accent_hover'], corner_radius=6,
                      font=ctk.CTkFont(size=13, weight="bold"),
                      command=on_ok).pack(side="left", padx=10)
        ctk.CTkButton(btn_frame, text="取消", fg_color=c['card_border'],
                      hover_color=c['sidebar_hover'], corner_radius=6,
                      font=ctk.CTkFont(size=13, weight="bold"),
                      command=on_cancel).pack(side="right", padx=10)

        dialog.wait_window()
        return self.preview_extra_columns

    def choose_filename_columns(self):
        c = self.colors
        dialog = ctk.CTkToplevel(self)
        dialog.title("选择文件名组成列")
        dialog.geometry("500x600")
        dialog.transient(self)
        dialog.grab_set()
        dialog.configure(fg_color=c['main_bg'])

        columns = list(self.df.columns)
        default_cols = config.DEFAULT_FILENAME_COLUMNS.copy()
        # 行号是虚拟列，始终放在第一位
        row_num_var = ctk.BooleanVar(value=("行号" in default_cols))
        check_vars = {col: ctk.BooleanVar(value=(col in default_cols)) for col in columns}

        main_frame = ctk.CTkFrame(dialog, fg_color="transparent")
        main_frame.pack(fill="both", expand=True, padx=20, pady=20)

        btn_bar = ctk.CTkFrame(main_frame, fg_color="transparent")
        btn_bar.pack(fill="x", pady=(0, 10))

        def _all_true():
            row_num_var.set(True)
            for v in check_vars.values(): v.set(True)
        def _all_false():
            row_num_var.set(False)
            for v in check_vars.values(): v.set(False)

        ctk.CTkButton(btn_bar, text="全选", width=80,
                      fg_color=c['success'], hover_color=c['success_hover'],
                      font=ctk.CTkFont(size=13, weight="bold"), corner_radius=6,
                      command=_all_true).pack(side="left", padx=5)
        ctk.CTkButton(btn_bar, text="取消全选", width=80,
                      fg_color=c['danger'], hover_color=c['danger_hover'],
                      font=ctk.CTkFont(size=13, weight="bold"), corner_radius=6,
                      command=_all_false).pack(side="left", padx=5)

        scroll = ctk.CTkScrollableFrame(main_frame, label_text="请按顺序勾选要加入文件名的列（可拖动排序）",
                                         fg_color=c['card_bg'])
        scroll.pack(fill="both", expand=True)

        # 行号固定在第一位
        ctk.CTkCheckBox(scroll, text="【行号】（自动编号，推荐放在第一位）", variable=row_num_var,
                        text_color=c['accent'], font=ctk.CTkFont(size=14, weight="bold")).pack(anchor="w", pady=3)

        for col in columns:
            ctk.CTkCheckBox(scroll, text=col, variable=check_vars[col],
                            text_color=c['text_primary']).pack(anchor="w", pady=2)

        # === 文件名实时预览 ===
        preview_frame = ctk.CTkFrame(main_frame, fg_color=c['accent_light'], corner_radius=6)
        preview_frame.pack(fill="x", pady=(6, 0))

        ctk.CTkLabel(preview_frame, text="预览：",
                     font=ctk.CTkFont(size=13, weight="bold"),
                     text_color=c['accent']).pack(side="left", padx=10, pady=8)

        preview_text = ctk.CTkLabel(preview_frame, text="（请勾选列名）",
                                    font=ctk.CTkFont(size=13),
                                    text_color=c['text_primary'],
                                    anchor="w", wraplength=400)
        preview_text.pack(side="left", fill="x", expand=True, padx=(0, 10), pady=8)

        def update_preview(*_):
            # 构建示例文件名
            sample_row = self.df.iloc[0]
            row_idx = sample_row.name
            parts = []
            if row_num_var.get():
                parts.append(str(row_idx + 1))
            for col in columns:
                if check_vars[col].get():
                    val = str(sample_row.get(col, '')) if pd.notna(sample_row.get(col, '')) else ''
                    val = re.sub(r'[\\/*?:"<>|]', '_', val)
                    val = val.replace('\n', '').replace('\r', '').strip() or '空'
                    parts.append(val)
            if parts:
                name = '_'.join(parts) + '.docx'
                if len(name) > config.MAX_FILENAME_LENGTH:
                    name = name[:config.MAX_FILENAME_LENGTH] + '…'
                preview_text.configure(text=name)
            else:
                preview_text.configure(text="（请至少勾选一列）")

        # 监听所有 checkbox 变化
        row_num_var.trace_add("write", update_preview)
        for var in check_vars.values():
            var.trace_add("write", update_preview)
        update_preview()

        selected_cols = []
        use_row_num = [False]

        def on_ok():
            nonlocal selected_cols
            use_row_num[0] = row_num_var.get()
            selected_cols = [col for col in columns if check_vars[col].get()]
            dialog.destroy()

        btn_frame = ctk.CTkFrame(dialog, fg_color="transparent")
        btn_frame.pack(fill="x", padx=20, pady=10)
        ctk.CTkButton(btn_frame, text="确定", fg_color=c['accent'],
                      hover_color=c['accent_hover'], corner_radius=6,
                      font=ctk.CTkFont(size=13, weight="bold"),
                      command=on_ok).pack(side="left", padx=10)
        ctk.CTkButton(btn_frame, text="取消", fg_color=c['accent'],
                      hover_color=c['accent_hover'], text_color="#ffffff", corner_radius=6,
                      command=dialog.destroy).pack(side="right", padx=10)

        dialog.wait_window()
        if use_row_num[0] and "行号" not in selected_cols:
            selected_cols.insert(0, "行号")
        return selected_cols

    def show_about(self):
        c = self.colors
        about_window = ctk.CTkToplevel(self)
        about_window.title("关于")
        about_window.geometry("460x520" if LITE_MODE else "460x380")
        about_window.resizable(False, False)
        about_window.grab_set()
        about_window.configure(fg_color=c['main_bg'])

        main_frame = ctk.CTkFrame(about_window, corner_radius=12,
                                  fg_color=c['card_bg'], border_color=c['card_border'], border_width=1)
        main_frame.pack(fill="both", expand=True, padx=20, pady=20)

        ctk.CTkLabel(main_frame, text="智汇文枢 Lite",
                     font=ctk.CTkFont(family="Microsoft YaHei", size=28, weight="bold"),
                     text_color=c['accent']).pack(pady=(24, 6))

        ctk.CTkFrame(main_frame, height=1, fg_color=c['card_border']).pack(fill="x", padx=40, pady=4)

        ctk.CTkLabel(main_frame,
                     text="高效的文档自动化工具，根据 Excel 数据\n自动填充 Word 模板生成文书。",
                     font=ctk.CTkFont(size=14), text_color=c['text_secondary'],
                     justify="center").pack(pady=(8, 2))

        ctk.CTkLabel(main_frame, text="作者：sam荒野",
                     font=ctk.CTkFont(family="Microsoft YaHei", size=14, weight="bold"),
                     text_color=c['text_primary']).pack(pady=(2, 0))

        # 定制引导（精简版）
        if LITE_MODE:
            ctk.CTkFrame(main_frame, height=1, fg_color=c['card_border']).pack(fill="x", padx=40, pady=8)
            ctk.CTkLabel(main_frame,
                         text="使用中有任何问题，欢迎加微信咨询，作者在线解答。\n\n"
                              "如需深度定制，例如：\n"
                              "  ▸ 内置物业费、违约金等自动计算公式\n"
                              "  ▸ 自定义公式编辑器\n"
                              "  ▸ 特定法院 / 行业文书模板适配\n"
                              "欢迎联系作者进行定制开发",
                         font=ctk.CTkFont(family="Microsoft YaHei", size=13),
                         text_color=c['text_secondary'], justify="left",
                         ).pack(padx=30, pady=(0, 2))
            ctk.CTkLabel(main_frame, text="微信: Sylvan_33ovo  |  QQ: 1784692583",
                         font=ctk.CTkFont(size=13, weight="bold"),
                         text_color=c['accent']).pack(pady=(0, 8))

        ctk.CTkLabel(main_frame, text="© 2026 sam荒野",
                     font=ctk.CTkFont(size=11), text_color=c['sidebar_text']).pack(pady=(8, 0))

        ctk.CTkButton(main_frame, text="关闭", width=120, height=36,
                      fg_color=c['accent'], hover_color=c['accent_hover'],
                      corner_radius=6,
                      font=ctk.CTkFont(size=15, weight="bold"),
                      command=about_window.destroy).pack(pady=(12, 16))

    def show_tutorial(self):
        c = self.colors
        tutorial_window = ctk.CTkToplevel(self)
        tutorial_window.title("使用教程")
        tutorial_window.geometry("680x480")
        tutorial_window.resizable(False, False)
        tutorial_window.grab_set()
        tutorial_window.configure(fg_color=c['main_bg'])

        main_frame = ctk.CTkFrame(tutorial_window, corner_radius=12,
                                  fg_color=c['card_bg'], border_color=c['card_border'], border_width=1)
        main_frame.pack(fill="both", expand=True, padx=20, pady=20)

        # 标题条
        title_bar = ctk.CTkFrame(main_frame, fg_color=c['header_bg'], corner_radius=8)
        title_bar.pack(fill="x", padx=16, pady=(14, 6))
        ctk.CTkLabel(title_bar, text="智汇文枢 Lite · 使用教程",
                     font=ctk.CTkFont(family="Microsoft YaHei", size=24, weight="bold"),
                     text_color=c['accent']).pack(pady=14)

        scroll_frame = ctk.CTkScrollableFrame(main_frame, fg_color="transparent",
                                               scrollbar_fg_color=c['card_border'])
        scroll_frame.pack(fill="both", expand=True, padx=16, pady=6)

        sections = [
            ("📄 第一步：准备 Word 模板（关键步骤）",
             "1. 打开您现有的 Word 文档（如物业费诉状模板）。\n"
             "2. 找到文档中需要替换的动态内容，用 {{ }} 双花括号包裹起来。\n"
             "   例如：\n"
             "     原文 → 张三   改成 → {{业主姓名}}\n"
             "     原文 → 2024年1月至2025年12月   改成 → {{物业服务费欠费周期}}\n"
             "     原文 → 5000.00   改成 → {{标的总额}}\n"
             "3. 所有需要从 Excel 填入的内容都要加上 {{ }}，名称可以自定义。\n"
             "4. 保存为 .docx 格式，这就是您的模板文件。\n\n"
             "💡 技巧：占位符名称建议使用中文，方便与 Excel 列名对应。\n"
             "   常见的占位符：{{业主姓名}}、{{身份证地址}}、{{联系电话}}、{{欠费合计}} 等。"),
            ("📊 第二步：准备 Excel 数据",
             "1. 准备一个 .xlsx 文件，第一行为表头（列名）。\n"
             "2. 列名需要与 Word 模板中的 {{占位符}} 名称一致，软件会自动匹配。\n"
             "   例如：模板中有 {{业主姓名}}，Excel 中就要有「业主姓名」这一列。\n"
             "3. 以下列名会被自动识别并智能处理：\n"
             "   业主姓名 → 支持顿号（、）拆分多个被申请人，如：张三、李四\n"
             "   身份证号 → 支持两种方式录入多个身份证号：\n"
             "     方式一：单列用顿号分隔，如 3303...、3303...（推荐）\n"
             "     方式二：多列分别录入，如「身份证号一」「身份证号二」\n"
             "   身份证地址 → 支持顿号拆分；如第二人是同地址，可填写\n"
             "     「同地址」「地址同上」「地址同前」等，软件自动复用第一人地址\n"
             "4. ⚠️ 身份证号列请务必设为「文本」格式，避免被 Excel 转成科学记数法！\n"
             "   （选中列 → 右键 → 设置单元格格式 → 文本 → 保存）"),
            ("🔍 第三步：处理原始数据",
             "1. 点击主界面「🔧 处理原始数据」按钮。\n"
             "2. 选择处理后文件的保存位置。\n"
             "3. 系统会自动比对 Word 占位符与 Excel 列，弹出变量校验报告：\n"
             "   ● 绿色 = 已匹配，数据会正确填入\n"
             "   ○ 红色 = 缺失，模板中该占位符暂无数据来源\n"
             "4. 确认后系统开始处理：\n"
             "   自动拆分姓名（张三、李四 → 被申请人一/二）\n"
             "   自动拆分身份证地址\n"
             "   自动解析身份证号（→ 出生年月日 + 性别）\n"
             "5. 处理完成后，预览区自动加载数据。"),
            ("📋 第四步：预览 & 勾选",
             "1. 预览区显示每条记录的关键信息（序号 + 业主姓名）。\n"
             "2. 点击「📋 选择预览列」可自定义显示更多列（如身份证地址等）。\n"
             "3. 勾选需要生成文书的条目（支持一键「✅ 全选」/「❌ 全不选」）。\n"
             "4. 点击「👁 单份预览」可先生成一份检查效果，确认无误再批量生成。"),
            ("📝 第五步：批量生成 Word 文档",
             "1. 勾选好要生成的条目后，点击「📝 生成目标文档」。\n"
             "2. 选择文件名的组成规则（推荐：行号 + 业主姓名，便于排序查阅）。\n"
             "3. 选择 Word 文件的保存文件夹。\n"
             "4. 系统批量生成，每份文档对应一条 Excel 记录。\n\n"
             "💡 建议先「单份预览」检查内容，确认无误后再批量生成。"),
        ]

        for title, body in sections:
            ctk.CTkLabel(scroll_frame, text=title,
                         font=ctk.CTkFont(family="Microsoft YaHei", size=17, weight="bold"),
                         text_color=c['accent']).pack(anchor="w", padx=4, pady=(12, 2))
            ctk.CTkLabel(scroll_frame, text=body,
                         font=ctk.CTkFont(family="Microsoft YaHei", size=14),
                         text_color=c['text_primary'],
                         justify="left", wraplength=600).pack(anchor="w", padx=12, pady=(0, 8))

        # 底部提示
        tip_bar = ctk.CTkFrame(main_frame, fg_color=c['accent_light'], corner_radius=6)
        tip_bar.pack(fill="x", padx=16, pady=(2, 8))
        ctk.CTkLabel(tip_bar, text="💡 身份证号默认识别「身份证号」列，支持顿号拆分  |  列名与占位符一致即可自动匹配  |  支持自定义文件名组成规则",
                     font=ctk.CTkFont(family="Microsoft YaHei", size=13),
                     text_color=c['accent'], wraplength=620).pack(padx=14, pady=10)

        ctk.CTkButton(main_frame, text="知道了", width=120, height=36,
                      fg_color=c['accent'], hover_color=c['accent_hover'],
                      corner_radius=6,
                      font=ctk.CTkFont(size=16, weight="bold"),
                      command=tutorial_window.destroy).pack(pady=(6, 16))

    def _build_main(self):
        """构建主内容区域（侧边栏右侧）"""
        c = self.colors
        parent = self.main_area

        # ===== 顶部标题栏 =====
        header = ctk.CTkFrame(parent, fg_color=c['header_bg'], corner_radius=8)
        header.grid(row=0, column=0, padx=20, pady=(16, 8), sticky="ew")
        ctk.CTkLabel(
            header, text="智汇文枢",
            font=ctk.CTkFont(family="Microsoft YaHei", size=28, weight="bold"),
            text_color=c['text_primary'],
        ).pack(side="left", padx=18, pady=14)

        # ===== 文件选择卡片 =====
        self._file_card = ctk.CTkFrame(
            parent, corner_radius=10, border_width=1,
            fg_color=c['card_bg'], border_color=c['card_border'],
        )
        self._file_card.grid(row=1, column=0, padx=20, pady=6, sticky="ew")
        self._file_card.grid_columnconfigure(1, weight=1)

        # 卡片标题 — 带左侧装饰线 + 微亮底
        title_bg = c['header_bg'] if self._dark_mode else c['accent_light']
        title_row = ctk.CTkFrame(self._file_card, fg_color=title_bg, corner_radius=6)
        title_row.grid(row=0, column=0, columnspan=3, padx=12, pady=(10, 4), sticky="ew")
        ctk.CTkFrame(title_row, width=3, height=18, fg_color=c['accent'], corner_radius=2).pack(side="left", padx=(10, 8))
        ctk.CTkLabel(
            title_row, text="文件选择",
            font=ctk.CTkFont(size=17, weight="bold"),
            text_color=c['text_primary'],
        ).pack(side="left", pady=5)

        # 文件行
        rows = [
            ("Word 模板：", "word_entry", self.select_word),
            ("原始 Excel：", "excel_entry", self.select_original_excel),
            ("处理后 Excel：", "processed_entry", self.select_processed_excel),
            ("输出文件夹：", "folder_entry", self.select_output_folder),
        ]
        for i, (label, attr, cmd) in enumerate(rows):
            ctk.CTkLabel(
                self._file_card, text=label,
                font=ctk.CTkFont(size=15, weight="bold"), text_color=c['text_secondary'],
            ).grid(row=i + 1, column=0, padx=(16, 4), pady=5, sticky="w")
            entry = ctk.CTkEntry(
                self._file_card, height=36,
                font=ctk.CTkFont(size=14),
                fg_color=c['entry_bg'], border_color=c['entry_border'],
                corner_radius=6,
            )
            entry.grid(row=i + 1, column=1, padx=4, pady=5, sticky="ew")
            setattr(self, attr, entry)
            ctk.CTkButton(
                self._file_card, text="浏览", width=60, height=32,
                fg_color=c['sidebar_hover'], hover_color=c['accent'],
                text_color="#ffffff", corner_radius=6,
                font=ctk.CTkFont(size=14, weight="bold"),
                command=cmd,
            ).grid(row=i + 1, column=2, padx=(4, 16), pady=5)

        # ===== 模板类型卡片 =====
        self._template_card = ctk.CTkFrame(
            parent, corner_radius=10, border_width=1,
            fg_color=c['card_bg'], border_color=c['card_border'],
        )
        self._template_card.grid(row=2, column=0, padx=20, pady=6, sticky="ew")
        self._template_card.grid_columnconfigure(1, weight=1)

        tpl_title = ctk.CTkFrame(self._template_card, fg_color=title_bg, corner_radius=6)
        tpl_title.grid(row=0, column=0, columnspan=3, padx=12, pady=(10, 4), sticky="ew")
        ctk.CTkFrame(tpl_title, width=3, height=18, fg_color=c['accent'], corner_radius=2).pack(side="left", padx=(10, 8))
        ctk.CTkLabel(
            tpl_title, text="模板设置",
            font=ctk.CTkFont(size=17, weight="bold"), text_color=c['text_primary'],
        ).pack(side="left", pady=5)

        ctk.CTkLabel(
            self._template_card, text="模板类型：",
            font=ctk.CTkFont(size=15, weight="bold"), text_color=c['text_secondary'],
        ).grid(row=1, column=0, padx=(16, 4), pady=6, sticky="w")

        self.template_menu = ctk.CTkOptionMenu(
            self._template_card, values=config.TEMPLATE_TYPES,
            variable=self.template_type, width=150,
            fg_color=c['entry_bg'], button_color=c['accent'],
            button_hover_color=c['accent_hover'],
            text_color=c['text_primary'],
        )
        self.template_menu.grid(row=1, column=1, padx=4, pady=6, sticky="w")

        if LITE_MODE:
            self.template_menu.configure(state="disabled")

        if not LITE_MODE:
            self.formula_btn = ctk.CTkButton(
                self._template_card, text="公式编辑", width=80, height=32,
                fg_color=c['sidebar_hover'], hover_color=c['accent'], text_color="#ffffff",
                font=ctk.CTkFont(size=13, weight="bold"), corner_radius=6,
                command=self.open_formula_editor, state="disabled",
            )
            self.formula_btn.grid(row=1, column=2, padx=(8, 16), pady=6, sticky="e")

        # 模板类型说明标签
        template_hints = {
            "预设公式型（瓯海）": "内置瓯海法院物业费/违约金/车位费/水费全套公式，自动计算",
            "预设公式型（平阳）": "与瓯海结构相似，违约金三段式计算参数略有差异",
            "通用填充型": "不自动计算，纯文本映射，自动提取身份证信息",
            "自定义公式型": "完全由用户编写计算公式，灵活度最高，适用任意模板",
        }
        self._template_hint_label = ctk.CTkLabel(
            self._template_card, text=template_hints[self.template_type.get()],
            font=ctk.CTkFont(size=12), text_color=c['text_secondary'],
        )
        self._template_hint_label.grid(row=2, column=0, columnspan=3, padx=(28, 16), pady=(0, 8), sticky="w")

        def on_template_change(*args):
            t = self.template_type.get()
            if not LITE_MODE:
                is_custom = t == "自定义公式型"
                self.formula_btn.configure(state="normal" if is_custom else "disabled")
            self._template_hint_label.configure(text=template_hints.get(t, ""))
        self.template_type.trace_add("write", on_template_change)

        # ===== 操作按钮 =====
        self._action_card = ctk.CTkFrame(
            parent, corner_radius=10, border_width=1,
            fg_color=c['card_bg'], border_color=c['card_border'],
        )
        self._action_card.grid(row=3, column=0, padx=20, pady=6, sticky="ew")
        self._action_card.grid_columnconfigure((0, 1, 2), weight=1)

        self.process_btn = ctk.CTkButton(
            self._action_card, text="处理原始数据", height=40,
            fg_color=c['accent'], hover_color=c['accent_hover'],
            font=ctk.CTkFont(size=14, weight="bold"), corner_radius=6,
            command=self.process_excel,
        )
        self.process_btn.grid(row=0, column=0, padx=12, pady=14, sticky="ew")

        self.generate_btn = ctk.CTkButton(
            self._action_card, text="生成目标文档", height=40,
            fg_color=c['success'], hover_color=c['success_hover'],
            font=ctk.CTkFont(size=14, weight="bold"), corner_radius=6,
            state="disabled", command=self.generate_selected,
        )

        self.preview_btn = ctk.CTkButton(
            self._action_card, text="单份预览", height=40,
            fg_color=c['warning'], hover_color=c['sidebar_hover'],
            text_color=c['text_primary'],
            font=ctk.CTkFont(size=14, weight="bold"), corner_radius=6,
            state="disabled", command=self.preview_single,
        )
        self.generate_btn.grid(row=0, column=1, padx=12, pady=14, sticky="ew")
        self.preview_btn.grid(row=0, column=2, padx=12, pady=14, sticky="ew")

        # ===== 状态标签 =====
        self.status_label = ctk.CTkLabel(
            parent, text="请先选择 Word 模板和原始 Excel 文件",
            font=ctk.CTkFont(size=14, weight="bold"), text_color=c['text_secondary'],
        )
        self.status_label.grid(row=4, column=0, padx=24, pady=(0, 2), sticky="w")

        # ===== 预览控制 =====
        self._ctrl_card = ctk.CTkFrame(
            parent, corner_radius=10, border_width=1,
            fg_color=c['card_bg'], border_color=c['card_border'],
        )
        self._ctrl_card.grid(row=5, column=0, padx=20, pady=(2, 0), sticky="ew")

        ctk.CTkButton(
            self._ctrl_card, text="全选", width=80, height=30,
            fg_color=c['success'], hover_color=c['success_hover'],
            font=ctk.CTkFont(size=13, weight="bold"), corner_radius=6,
            command=self.select_all,
        ).pack(side="left", padx=12, pady=8)
        ctk.CTkButton(
            self._ctrl_card, text="取消全选", width=80, height=30,
            fg_color=c['danger'], hover_color=c['danger_hover'],
            font=ctk.CTkFont(size=13, weight="bold"), corner_radius=6,
            command=self.clear_all,
        ).pack(side="left", padx=4, pady=8)
        ctk.CTkButton(
            self._ctrl_card, text="选择预览列", width=100, height=30,
            fg_color=c['sidebar_hover'], hover_color=c['accent'], text_color="#ffffff",
            font=ctk.CTkFont(size=13, weight="bold"), corner_radius=6,
            command=self.choose_preview_columns,
        ).pack(side="right", padx=12, pady=8)

        # ===== 预览区 =====
        self._preview_card = ctk.CTkFrame(
            parent, corner_radius=10, border_width=1,
            fg_color=c['card_bg'], border_color=c['card_border'],
        )
        self._preview_card.grid(row=6, column=0, padx=20, pady=6, sticky="nsew")
        self._preview_card.grid_rowconfigure(1, weight=1)
        self._preview_card.grid_columnconfigure(0, weight=1)

        pv_header_bg = c['header_bg'] if self._dark_mode else c['accent_light']
        pv_title = ctk.CTkFrame(self._preview_card, fg_color=pv_header_bg, corner_radius=6)
        pv_title.grid(row=0, column=0, padx=12, pady=(10, 2), sticky="ew")
        ctk.CTkFrame(pv_title, width=3, height=16, fg_color=c['accent'], corner_radius=2).pack(side="left", padx=(10, 8))
        ctk.CTkLabel(
            pv_title, text="数据预览（勾选要生成的条目）",
            font=ctk.CTkFont(size=14, weight="bold"), text_color=c['text_primary'],
        ).pack(side="left", pady=6)

        self.preview_scroll = TwoWayScrollableFrame(
            self._preview_card, corner_radius=6,
            fg_color=c['card_bg'],
        )
        # 强制 Canvas 和内层 Frame 底色统一
        self.preview_scroll.canvas.configure(bg=c['card_bg'])
        self.preview_scroll.canvas_bg = c['card_bg']
        self.preview_scroll.inner_frame.configure(fg_color=c['card_bg'])
        self.preview_scroll.grid(row=1, column=0, padx=12, pady=(2, 12), sticky="nsew")

        # ===== 进度条 =====
        self.progress_frame = ctk.CTkFrame(parent, fg_color="transparent")
        self.progress_frame.grid(row=7, column=0, padx=20, pady=(2, 8), sticky="ew")
        self.progress_frame.grid_columnconfigure(0, weight=1)

        self.progressbar = ctk.CTkProgressBar(
            self.progress_frame, height=6, corner_radius=3,
            progress_color=c['accent'],
            fg_color=c['card_border'],
        )
        self.progressbar.grid(row=0, column=0, padx=(0, 8), pady=4, sticky="ew")
        self.progressbar.set(0)

        self.progress_label = ctk.CTkLabel(
            self.progress_frame, text="0%",
            font=ctk.CTkFont(size=14, weight="bold"), text_color=c['text_secondary'],
        )
        self.progress_label.grid(row=0, column=1, padx=4, pady=4)

        self.cancel_btn = ctk.CTkButton(
            self.progress_frame, text="取消", width=60, height=28,
            fg_color=c['danger'], hover_color=c['danger_hover'],
            font=ctk.CTkFont(size=13, weight="bold"), command=self.cancel_task, state="disabled",
        )
        self.cancel_btn.grid(row=0, column=2, padx=8, pady=4)
        self.progress_frame.grid_remove()

        self.preview_inner = self.preview_scroll.get_inner_frame()
        self.status_label.configure(wraplength=800)

    def select_word(self):
        path = filedialog.askopenfilename(title="选择 Word 模板", filetypes=[("Word 文件", "*.docx")])
        if path:
            self.word_path = path
            self.word_entry.delete(0, "end")
            self.word_entry.insert(0, path)
            try:
                self.placeholders = extract_placeholders_from_template(path)
            except Exception as e:
                self.placeholders = []
                messagebox.showwarning("警告", f"无法解析模板占位符，生成时可能出错：{str(e)}")
            self.update_status()

    def select_original_excel(self):
        path = filedialog.askopenfilename(title="选择原始 Excel", filetypes=[("Excel 文件", "*.xlsx")])
        if path:
            self.original_excel_path = path
            self.excel_entry.delete(0, "end")
            self.excel_entry.insert(0, path)
            self.update_status()

    def update_status(self):
        c = self.colors
        if self.word_path and self.original_excel_path:
            self.status_label.configure(
                text="● 已就绪，请点击「处理原始数据」",
                text_color=c['success'],
            )
        else:
            self.status_label.configure(
                text="● 请先选择 Word 模板和原始 Excel 文件",
                text_color=c['text_secondary'],
            )

    def select_processed_excel(self):
        path = filedialog.askopenfilename(title="选择已处理好的 Excel", filetypes=[("Excel 文件", "*.xlsx")])
        if path:
            self.processed_excel_path = path
            self.processed_entry.delete(0, "end")
            self.processed_entry.insert(0, path)
            self.load_processed_excel(path)

    def load_processed_excel(self, path):
        try:
            self.check_vars.clear()
            for widget in self.preview_inner.winfo_children():
                widget.destroy()

            df = pd.read_excel(path).fillna('')
            # 身份证/电话列强制文本，防止科学记数法
            for col in df.columns:
                cstr = str(col)
                if '身份证' in cstr or '电话' in cstr:
                    df[col] = df[col].apply(lambda x: '' if pd.isna(x) else (f'{x:.0f}' if isinstance(x, float) else str(x)))
            required_cols = ['序号', '身份证地址', '业主姓名']
            missing = [col for col in required_cols if col not in df.columns]
            if missing:
                messagebox.showwarning("列缺失", f"该 Excel 缺少以下列，可能不是有效的处理后文件：\n{missing}")
            self.df = df

            if self.word_path and os.path.exists(self.word_path):
                try:
                    self.placeholders = extract_placeholders_from_template(self.word_path)
                except Exception as e:
                    self.placeholders = []
                    messagebox.showwarning("警告", f"无法解析模板占位符：{str(e)}")
            else:
                self.placeholders = []

            self.load_preview()
            self.generate_btn.configure(state="normal")
            self.preview_btn.configure(state="normal")
            self.status_label.configure(text=f"已加载处理后 Excel：{os.path.basename(path)}，共 {len(df)} 条记录")
        except Exception as e:
            messagebox.showerror("加载失败", f"无法加载 Excel 文件：\n{str(e)}")

    def select_output_folder(self):
        path = filedialog.askdirectory(title="选择 Word 输出文件夹")
        if path:
            self.output_folder = path
            self.folder_entry.delete(0, "end")
            self.folder_entry.insert(0, path)

    def process_excel(self):
        if not self.original_excel_path:
            messagebox.showwarning("提示", "请先选择原始 Excel 文件")
            return

        base_name = os.path.splitext(os.path.basename(self.original_excel_path))[0]
        default_filename = f"{base_name}_处理后.xlsx"
        initial_dir = os.path.dirname(self.original_excel_path)
        save_path = filedialog.asksaveasfilename(
            title="保存处理后的 Excel 文件",
            initialdir=initial_dir,
            initialfile=default_filename,
            defaultextension=".xlsx",
            filetypes=[("Excel 文件", "*.xlsx")]
        )
        if not save_path:
            if not messagebox.askyesno("保存取消", "是否使用默认路径保存？"):
                self.status_label.configure(text="处理已取消")
                return
            else:
                save_path = os.path.join(initial_dir, default_filename)

        self.processed_excel_path = save_path
        self.processed_entry.delete(0, "end")
        self.processed_entry.insert(0, save_path)

        # === 处理前变量校验 ===
        if not self.word_path or not os.path.exists(self.word_path):
            messagebox.showwarning("提示", "请先选择有效的 Word 模板文件")
            return

        # 提取 Word 占位符
        if not self.placeholders:
            self.placeholders = extract_placeholders_from_template(self.word_path)

        # 读取 Excel 列名（预读，不实际处理）
        try:
            preview_df = pd.read_excel(self.original_excel_path)
            excel_cols = list(preview_df.columns)
        except Exception as e:
            messagebox.showerror("错误", f"无法读取 Excel：{str(e)}")
            return

        # 公式结果列（仅自定义模板）
        formula_results = []
        if self.template_type.get() == "自定义公式型" and self.formulas:
            from models.formula_engine import parse_formula
            for f in self.formulas:
                p = parse_formula(f)
                if p:
                    formula_results.append(p[1])

        # 变量比对
        from models.formula_engine import match_variables
        all_cols = excel_cols + formula_results
        matched, unmatched, summary = match_variables(self.placeholders, all_cols, [])

        # 弹窗确认
        confirmed = [False]
        def on_confirm():
            confirmed[0] = True
            dlg.destroy()

        c = self.colors
        dlg = ctk.CTkToplevel(self)
        dlg.title("变量校验 — 处理前确认")
        dlg.geometry("520x500")
        dlg.transient(self)
        dlg.grab_set()
        dlg.configure(fg_color=c['main_bg'])

        sc = ctk.CTkFrame(dlg, fg_color=c['card_bg'], corner_radius=8, border_width=1, border_color=c['card_border'])
        sc.pack(fill="x", padx=16, pady=(16, 8))
        ctk.CTkLabel(sc, text=f"模板需要 {summary['total']} 个变量  |  已匹配 {summary['matched_count']} 个  |  缺失 {summary['unmatched_count']} 个",
                     font=ctk.CTkFont(size=15, weight="bold"),
                     text_color=c['success'] if summary['unmatched_count'] == 0 else c['danger']).pack(padx=16, pady=14)

        scroll = ctk.CTkScrollableFrame(dlg, fg_color=c['card_bg'], corner_radius=8, border_width=1, border_color=c['card_border'])
        scroll.pack(fill="both", expand=True, padx=16, pady=6)

        if matched:
            ctk.CTkLabel(scroll, text="已匹配",
                         font=ctk.CTkFont(size=14, weight="bold"), text_color=c['success']).pack(anchor="w", padx=14, pady=(10, 2))
            for vn, src in matched:
                ctk.CTkLabel(scroll, text=f"  {vn}",
                             font=ctk.CTkFont(size=14), text_color=c['success']).pack(anchor="w", padx=24)

        if unmatched:
            ctk.CTkLabel(scroll, text="缺失（将创建空列）",
                         font=ctk.CTkFont(size=14, weight="bold"), text_color=c['danger']).pack(anchor="w", padx=14, pady=(14, 2))
            for vn in unmatched:
                ctk.CTkLabel(scroll, text=f"  {vn}",
                             font=ctk.CTkFont(size=14), text_color=c['danger']).pack(anchor="w", padx=24)

        btn_bar = ctk.CTkFrame(dlg, fg_color="transparent")
        btn_bar.pack(fill="x", padx=16, pady=(8, 16))
        ctk.CTkButton(btn_bar, text="开始处理", height=36,
                      fg_color=c['success'], hover_color=c['success_hover'],
                      text_color="#ffffff", corner_radius=6,
                      font=ctk.CTkFont(size=15, weight="bold"),
                      command=on_confirm).pack(side="left", padx=6)
        ctk.CTkButton(btn_bar, text="取消", height=36,
                      fg_color=c['danger'], hover_color=c['danger_hover'],
                      text_color="#ffffff", corner_radius=6,
                      font=ctk.CTkFont(size=15, weight="bold"),
                      command=dlg.destroy).pack(side="right", padx=6)

        dlg.wait_window()

        if not confirmed[0]:
            self.status_label.configure(text="已取消 — 请补全数据后重试")
            return

        # === 身份证处理弹窗（仅通用/自定义模板） ===
        template = self.template_type.get()
        if template in ["通用填充型", "自定义公式型"]:
            dlg, id_result = self._show_id_processing_dialog()
            dlg.wait_window()
            if id_result[0] is None:
                self.status_label.configure(text="已取消")
                return
            id_config = id_result[0]
        else:
            # 瓯海/平阳模板使用默认配置（自动解析 身份证号一/二）
            id_config = {'enabled': True, 'split_enabled': False,
                        'source_column': '', 'manual_columns': []}

        # === 用户确认，开始实际处理 ===
        self.process_btn.configure(state="disabled")
        self.generate_btn.configure(state="disabled")
        self.preview_btn.configure(state="disabled")
        self.progress_frame.grid()
        self.progressbar.set(0)
        self.progress_label.configure(text="0%")
        self.cancel_requested = False
        self.cancel_btn.configure(state="normal")

        threading.Thread(target=self._process_excel_worker, args=(save_path, id_config), daemon=True).start()

    def load_preview(self):
        if not hasattr(self, 'preview_inner') or self.preview_inner is None:
            messagebox.showerror("错误", "预览区域未初始化，请重启程序")
            return

        self.preview_extra_columns = [col for col in self.preview_extra_columns if col in self.df.columns]
        if not self.preview_extra_columns:
            self.preview_extra_columns = ['身份证地址'] if '身份证地址' in self.df.columns else []

        if hasattr(self, 'progress_frame'):
            self.progress_frame.grid_remove()

        for widget in self.preview_inner.winfo_children():
            widget.destroy()

        if self.df is None or self.df.empty:
            ctk.CTkLabel(self.preview_inner, text="暂无数据",
                         font=ctk.CTkFont(size=14),
                         text_color=c['text_secondary'],
                         fg_color="transparent").pack(pady=50)
            return

        fixed_columns = ['选择', '序号', '业主姓名']
        display_columns = fixed_columns + self.preview_extra_columns
        col_widths = [30, 50, 200] + [150] * len(self.preview_extra_columns)
        c = self.colors
        # 深色: 极微妙交替色; 浅色: 冷灰
        alt_bg = c['sidebar_hover'] if self._dark_mode else "#F8FAFC"
        text_c = c['text_primary']

        for col_idx in range(len(display_columns)):
            if col_idx < 3:
                self.preview_inner.grid_columnconfigure(col_idx, weight=0, minsize=col_widths[col_idx])
            else:
                self.preview_inner.grid_columnconfigure(col_idx, weight=0, minsize=col_widths[col_idx])

        for col_idx, col_name in enumerate(display_columns):
            if col_idx < 3:
                width = col_widths[col_idx]
            else:
                width = col_widths[col_idx] if col_idx < len(col_widths) else 150
            header = ctk.CTkLabel(
                self.preview_inner,
                text=col_name,
                width=width,
                anchor="center",
                font=ctk.CTkFont(size=12, weight="bold"),
                corner_radius=4,
                text_color=c['accent'],
                fg_color=c['accent_light'],
            )
            header.grid(row=0, column=col_idx, padx=1, pady=2, sticky="ew")

        self.check_vars = []
        for idx, row in self.df.iterrows():
            row_index = idx + 1
            is_odd = idx % 2 == 1
            bg = alt_bg if is_odd else "transparent"
            bg_frame = ctk.CTkFrame(self.preview_inner, fg_color=bg, height=32)
            bg_frame.grid(row=row_index, column=0, columnspan=len(display_columns), sticky="nsew", pady=1)
            bg_frame.grid_propagate(False)
            bg_frame.lower()

            chk_container = ctk.CTkFrame(self.preview_inner, fg_color="transparent", height=32)
            chk_container.grid(row=row_index, column=0, sticky="n", pady=1)
            chk_container.columnconfigure(0, weight=1)
            chk_container.grid_propagate(False)

            var = ctk.BooleanVar(value=False)
            self.check_vars.append(var)
            chk = ctk.CTkCheckBox(chk_container, text="", variable=var, width=15)
            chk.grid(row=0, column=0, padx=2)

            lbl_seq = ctk.CTkLabel(
                self.preview_inner, text=str(row_index),
                width=col_widths[1], anchor="center",
                font=ctk.CTkFont(size=15), text_color=text_c,
                fg_color="transparent",
            )
            lbl_seq.grid(row=row_index, column=1, padx=1, pady=2, sticky="ew")

            name_value = str(row.get('业主姓名', ''))
            lbl_name = ctk.CTkLabel(
                self.preview_inner, text=name_value,
                width=col_widths[2], anchor="w",
                font=ctk.CTkFont(size=15), text_color=text_c,
                fg_color="transparent",
            )
            lbl_name.grid(row=row_index, column=2, padx=1, pady=2, sticky="ew")

            for col_idx, col_name in enumerate(self.preview_extra_columns, start=3):
                value = str(row.get(col_name, ''))
                lbl = ctk.CTkLabel(
                    self.preview_inner, text=value,
                    width=col_widths[col_idx] if col_idx < len(col_widths) else 150,
                    anchor="center", font=ctk.CTkFont(size=13), text_color=text_c,
                    fg_color="transparent",
                )
                lbl.grid(row=row_index, column=col_idx, padx=1, pady=2, sticky="ew")

        self.preview_inner.update_idletasks()
        bbox = self.preview_scroll.canvas.bbox("all")
        if bbox:
            self.preview_scroll.canvas.configure(scrollregion=bbox)
        self.preview_scroll.canvas.yview_moveto(0)

        def bind_mousewheel_recursive(widget):
            widget.bind("<MouseWheel>", self.preview_scroll._on_mousewheel, add="+")
            widget.bind("<Button-4>", self.preview_scroll._on_mousewheel_linux, add="+")
            widget.bind("<Button-5>", self.preview_scroll._on_mousewheel_linux, add="+")
            for child in widget.winfo_children():
                bind_mousewheel_recursive(child)

        bind_mousewheel_recursive(self.preview_inner)
        self.generate_btn.configure(state="normal")
        self.preview_btn.configure(state="normal")
        self.status_label.configure(text=f"已加载 {len(self.df)} 条记录，勾选后生成Word")

    def select_all(self):
        for var in self.check_vars:
            var.set(True)

    def clear_all(self):
        for var in self.check_vars:
            var.set(False)

    def preview_single(self):
        """生成第一份勾选文档的预览，打开供用户确认"""
        if self.df is None:
            messagebox.showwarning("提示", "请先处理 Excel 或加载处理后 Excel")
            return
        if not self.word_path or not os.path.exists(self.word_path):
            messagebox.showwarning("提示", "请先选择有效的 Word 模板")
            return

        selected = [i for i, var in enumerate(self.check_vars) if var.get()]
        if not selected:
            messagebox.showwarning("提示", "请至少勾选一条记录用于预览")
            return

        # 取第一个勾选的行
        idx = selected[0]
        row = self.df.iloc[idx]

        if not self.placeholders:
            self.placeholders = extract_placeholders_from_template(self.word_path)

        from services.word_generator import _build_context
        from docxtpl import DocxTemplate

        try:
            context = _build_context(row, self.placeholders)

            # 保存为临时预览文件
            preview_dir = os.path.dirname(self.original_excel_path) if self.original_excel_path else os.path.expanduser("~")
            preview_path = os.path.join(preview_dir, "_预览_请检查后删除.docx")
            doc = DocxTemplate(self.word_path)
            doc.render(context)
            doc.save(preview_path)

            # 打开预览文件
            os.startfile(preview_path)
            messagebox.showinfo("预览已打开",
                f"预览文件：\n{preview_path}\n\n"
                "请检查内容是否正确，确认无误后批量生成。\n"
                "预览文件可手动删除。")
        except Exception as e:
            messagebox.showerror("预览失败", f"生成预览时出错：\n{str(e)}")

    def generate_selected(self):
        if self.df is None:
            messagebox.showwarning("提示", "请先处理 Excel 或加载处理后 Excel")
            return

        selected = [i for i, var in enumerate(self.check_vars) if var.get()]
        if not selected:
            messagebox.showwarning("提示", "请至少勾选一条记录")
            return

        selected_columns = self.choose_filename_columns()
        if not selected_columns:
            if not messagebox.askyesno("提示", "您未选择任何列，是否使用默认规则（行号_业主姓名_身份证地址）？"):
                return
            selected_columns = config.DEFAULT_FILENAME_COLUMNS.copy()

        output_dir = self.output_folder
        if not output_dir:
            # 默认使用 Excel 所在目录
            default_dir = ""
            if self.processed_excel_path and os.path.exists(self.processed_excel_path):
                default_dir = os.path.dirname(self.processed_excel_path)
            elif self.original_excel_path and os.path.exists(self.original_excel_path):
                default_dir = os.path.dirname(self.original_excel_path)
            output_dir = filedialog.askdirectory(title="选择保存文件夹", initialdir=default_dir)
            if not output_dir:
                return
            self.output_folder = output_dir
            self.folder_entry.delete(0, "end")
            self.folder_entry.insert(0, output_dir)

        if not self.placeholders and self.word_path:
            self.placeholders = extract_placeholders_from_template(self.word_path)

        self.generate_btn.configure(state="disabled")
        self.preview_btn.configure(state="disabled")
        self.process_btn.configure(state="disabled")
        self.progress_frame.grid()
        self.progressbar.set(0)
        self.progress_label.configure(text=f"0/{len(selected)}")
        self.cancel_requested = False
        self.cancel_btn.configure(state="normal")

        threading.Thread(target=self._generate_worker, args=(selected, selected_columns, output_dir), daemon=True).start()

    def _generate_worker(self, selected, selected_columns, output_dir):
        total = len(selected)

        def progress_cb(current, total_count):
            progress = current / total_count
            self.progress_queue.put({'type': 'progress', 'value': progress, 'text': f"{current}/{total_count}"})

        def cancel_cb():
            return self.cancel_requested

        placeholders = self.placeholders if self.placeholders else []
        count, failed, was_cancelled, truncated = generate_documents(
            df=self.df,
            selected_indices=selected,
            word_template_path=self.word_path,
            placeholders=placeholders,
            output_dir=output_dir,
            filename_columns=selected_columns,
            progress_callback=progress_cb,
            cancel_checker=cancel_cb,
        )

        if was_cancelled:
            self.progress_queue.put({'type': 'cancelled', 'text': "生成已取消"})
            return

        self.progress_queue.put({
            'type': 'done',
            'text': f"成功生成 {count} 份文件\n保存位置：\n{output_dir}",
            'count': count,
            'output_dir': output_dir
        })
        if count > 0:
            self.progress_queue.put({'type': 'open_folder', 'path': output_dir})

        if truncated:
            detail = "\n".join([f"  序号{t[0]}: {t[1][:60]}... → {t[2]}" for t in truncated[:5]])
            if len(truncated) > 5:
                detail += f"\n  ... 共 {len(truncated)} 个文件名被截断"
            self.after(0, lambda: messagebox.showwarning(
                "文件名过长已截断",
                f"以下 {len(truncated)} 个文件名超过 {config.MAX_FILENAME_LENGTH} 字符上限，"
                f"已自动截断处理：\n\n{detail}"
            ))

        if failed:
            self.after(0, lambda: messagebox.showwarning(
                "部分失败",
                f"成功生成 {count} 份文件，失败 {len(failed)} 个\n"
                f"首个失败：序号 {failed[0][0]} - {failed[0][1]}"
            ))
        self.progress_queue.put({'type': 'enable_buttons'})

    def open_formula_editor(self):
        """自定义公式编辑器 — 左侧公式输入 + 右侧变量对照"""
        c = self.colors
        dialog = ctk.CTkToplevel(self)
        dialog.title("自定义公式编辑")
        dialog.geometry("820x620")
        dialog.resizable(True, True)
        dialog.grab_set()
        dialog.configure(fg_color=c['main_bg'])

        panes = ctk.CTkFrame(dialog, fg_color="transparent")
        panes.pack(fill="both", expand=True, padx=12, pady=12)
        panes.grid_columnconfigure(0, weight=3)
        panes.grid_columnconfigure(1, weight=2)

        left = ctk.CTkFrame(panes, fg_color=c['card_bg'], corner_radius=8, border_width=1, border_color=c['card_border'])
        left.grid(row=0, column=0, padx=(0, 6), sticky="nsew")
        left.grid_rowconfigure(2, weight=1)
        left.grid_columnconfigure(0, weight=1)

        ctk.CTkLabel(left, text="计算公式（一行一个）",
                     font=ctk.CTkFont(size=16, weight="bold"), text_color=c['text_primary']).pack(anchor="w", padx=14, pady=(12, 4))
        ctk.CTkLabel(left, text="{{变量}} 运算符 {{变量}} = {{结果}}",
                     font=ctk.CTkFont(size=14), text_color=c['accent']).pack(anchor="w", padx=14)
        ctk.CTkLabel(left, text="示例: {{单价}}*{{建筑面积}}*12/365*{{欠费天数}}={{物业费总额}}",
                     font=ctk.CTkFont(size=14), text_color=c['text_secondary']).pack(anchor="w", padx=14, pady=(0, 8))

        formula_text = ctk.CTkTextbox(
            left, font=ctk.CTkFont(family="Consolas", size=14),
            fg_color=c['entry_bg'], border_color=c['entry_border'],
            border_width=1, corner_radius=6, wrap="word",
        )
        formula_text.pack(fill="both", expand=True, padx=14, pady=(2, 4))
        if self.formulas:
            formula_text.insert("1.0", "\n".join(self.formulas))

        # ===== 数字小键盘 =====
        def insert_char(char):
            """在光标位置插入字符"""
            pos = formula_text.index("insert")
            formula_text.insert(pos, char)
            formula_text.focus_set()

        def clear_formula():
            """清空公式文本框"""
            formula_text.delete("1.0", "end")
            formula_text.focus_set()

        keypad = ctk.CTkFrame(left, fg_color="transparent")
        keypad.pack(fill="x", padx=14, pady=(0, 6))
        for i in range(4):
            keypad.grid_rowconfigure(i, weight=1)
        for j in range(6):
            keypad.grid_columnconfigure(j, weight=1)

        btn_small = {'width': 36, 'height': 30, 'font': ctk.CTkFont(size=14, weight="bold"),
                     'corner_radius': 4, 'border_width': 1}
        btn_normal = {**btn_small, 'fg_color': c['entry_bg'], 'hover_color': c['accent_light'],
                      'text_color': c['text_primary'], 'border_color': c['entry_border']}
        btn_op = {**btn_small, 'fg_color': c['header_bg'],
                  'hover_color': c['accent'], 'text_color': '#ffffff',
                  'border_color': c['header_bg']}
        btn_accent = {**btn_small, 'fg_color': c['accent'],
                      'hover_color': c['accent_hover'], 'text_color': '#ffffff',
                      'border_color': c['accent']}
        btn_danger = {**btn_small, 'fg_color': c['danger'],
                      'hover_color': c['danger_hover'], 'text_color': '#ffffff',
                      'border_color': c['danger']}

        # 键盘布局: (text, insert_value/action, row, col, colspan, style)
        keys = [
            ("7", "7", 0, 0, 1, btn_normal), ("8", "8", 0, 1, 1, btn_normal),
            ("9", "9", 0, 2, 1, btn_normal), ("+", "+", 0, 3, 1, btn_op),
            ("%", "%", 0, 4, 1, btn_op), ("{{}}", "{{}}", 0, 5, 1, btn_op),

            ("4", "4", 1, 0, 1, btn_normal), ("5", "5", 1, 1, 1, btn_normal),
            ("6", "6", 1, 2, 1, btn_normal), ("−", "-", 1, 3, 1, btn_op),
            ("(", "(", 1, 4, 1, btn_op), (")", ")", 1, 5, 1, btn_op),

            ("1", "1", 2, 0, 1, btn_normal), ("2", "2", 2, 1, 1, btn_normal),
            ("3", "3", 2, 2, 1, btn_normal), ("×", "*", 2, 3, 1, btn_op),
            (".", ".", 2, 4, 1, btn_op), ("=", "=", 2, 5, 1, btn_accent),

            ("0", "0", 3, 0, 1, btn_normal), ("清空", clear_formula, 3, 1, 2, btn_danger),
            ("÷", "/", 3, 3, 1, btn_op),
        ]

        for item in keys:
            text, action, row, col, colspan, style = item
            if callable(action):
                btn = ctk.CTkButton(keypad, text=text, command=action, **style)
            else:
                btn = ctk.CTkButton(keypad, text=text, command=lambda c=action: insert_char(c), **style)
            btn.grid(row=row, column=col, columnspan=colspan, padx=1, pady=1, sticky="nsew")

        # 记录原始内容用于关闭时比对
        original_content = formula_text.get("1.0", "end-1c").strip()

        right = ctk.CTkFrame(panes, fg_color=c['card_bg'], corner_radius=8, border_width=1, border_color=c['card_border'])
        right.grid(row=0, column=1, padx=(6, 0), sticky="nsew")
        right.grid_rowconfigure(2, weight=1)
        right.grid_columnconfigure(0, weight=1)

        ctk.CTkLabel(right, text="变量对照",
                     font=ctk.CTkFont(size=16, weight="bold"), text_color=c['text_primary']).pack(anchor="w", padx=14, pady=(12, 4))

        wv = self.placeholders if self.placeholders else []
        ec = []
        if self.original_excel_path and os.path.exists(self.original_excel_path):
            try:
                ec = list(pd.read_excel(self.original_excel_path).columns)
            except Exception:
                ec = []

        rs = ctk.CTkScrollableFrame(right, fg_color="transparent")
        rs.pack(fill="both", expand=True, padx=8, pady=4)

        # === 插入变量辅助函数 ===
        def insert_var(var_name):
            """在公式文本框光标处插入 {{变量名}}"""
            pos = formula_text.index("insert")
            formula_text.insert(pos, "{{" + var_name + "}}")
            formula_text.focus_set()

        def make_clickable(lbl, var_name, normal_color):
            """绑定 hover + 点击事件到变量标签"""
            lbl.configure(cursor="hand2")
            lbl.bind("<Button-1>", lambda e, v=var_name: insert_var(v))
            lbl.bind("<Enter>", lambda e: lbl.configure(text_color=c['accent']))
            lbl.bind("<Leave>", lambda e: lbl.configure(text_color=normal_color))

        ctk.CTkLabel(rs, text=f"Word 变量（{len(wv)} 个）— 点击可插入",
                     font=ctk.CTkFont(size=14, weight="bold"), text_color=c['accent']).pack(anchor="w", padx=6, pady=(6, 2))
        if wv:
            for v in wv:
                ok = v in ec
                color = c['success'] if ok else c['danger']
                lbl = ctk.CTkLabel(rs, text=f"{'●' if ok else '○'} {v}",
                                   font=ctk.CTkFont(size=14), text_color=color)
                lbl.pack(anchor="w", padx=16)
                make_clickable(lbl, v, color)
        else:
            ctk.CTkLabel(rs, text="(请先选择 Word 模板)",
                         font=ctk.CTkFont(size=14), text_color=c['text_secondary']).pack(anchor="w", padx=16)

        ctk.CTkLabel(rs, text=f"Excel 列（{len(ec)} 个）— 点击可插入",
                     font=ctk.CTkFont(size=14, weight="bold"), text_color=c['accent']).pack(anchor="w", padx=6, pady=(14, 2))
        if ec:
            for col in ec:
                matched = col in wv
                color = c['success'] if matched else c['text_secondary']
                prefix = "●" if matched else "○"
                lbl = ctk.CTkLabel(rs, text=f"  {col}",
                                   font=ctk.CTkFont(size=14), text_color=color)
                lbl.pack(anchor="w", padx=16)
                make_clickable(lbl, col, color)
        else:
            ctk.CTkLabel(rs, text="(请先选择 Excel 文件)",
                         font=ctk.CTkFont(size=14), text_color=c['text_secondary']).pack(anchor="w", padx=16)

        ctk.CTkLabel(rs, text="● = 已匹配  ○ = 缺失  |  点击变量名即可插入",
                     font=ctk.CTkFont(size=13), text_color=c['text_secondary']).pack(anchor="w", padx=6, pady=(10, 0))

        def save():
            nonlocal original_content
            text = formula_text.get("1.0", "end-1c").strip()
            lines = [f.strip() for f in text.split('\n') if f.strip()]

            if not lines:
                messagebox.showwarning("提示", "请至少输入一条公式")
                return

            # 构建有效变量集合
            valid_vars = set(wv) | set(ec)

            # 逐行校验
            errors = []
            for i, line in enumerate(lines):
                is_valid, err = validate_formula(line, valid_vars)
                if not is_valid:
                    errors.append(f"第{i + 1}行: {err}")

            if errors:
                messagebox.showwarning("公式格式错误",
                    "以下公式校验未通过，请修正后重试：\n\n" + "\n".join(errors))
                return

            self.formulas = lines
            original_content = text  # 保存后更新原始内容，避免关闭时误报
            messagebox.showinfo("已保存", f"已保存 {len(self.formulas)} 条公式，全部校验通过")
            dialog.destroy()

        def on_close():
            current = formula_text.get("1.0", "end-1c").strip()
            if current != original_content:
                if not messagebox.askyesno("未保存", "公式已修改但未保存，是否放弃修改？", parent=dialog):
                    return
            dialog.destroy()

        btn_frame = ctk.CTkFrame(dialog, fg_color="transparent")
        btn_frame.pack(fill="x", padx=16, pady=(4, 14))
        ctk.CTkButton(btn_frame, text="保存公式", height=36, fg_color=c['accent'],
                      hover_color=c['accent_hover'], corner_radius=6,
                      font=ctk.CTkFont(size=15, weight="bold"),
                      command=save).pack(side="left", padx=6)
        ctk.CTkButton(btn_frame, text="取消", height=36, fg_color=c['sidebar_hover'],
                      hover_color=c['accent'], text_color="#ffffff", corner_radius=6,
                      font=ctk.CTkFont(size=15, weight="bold"),
                      command=on_close).pack(side="right", padx=6)

        dialog.protocol("WM_DELETE_WINDOW", on_close)

    def show_variable_validation(self, matched, unmatched, summary):
        """变量验证结果弹窗 — 绿色=已匹配, 红色=未匹配"""
        c = self.colors
        dialog = ctk.CTkToplevel(self)
        dialog.title("变量验证结果")
        dialog.geometry("520x550")
        dialog.resizable(False, False)
        dialog.grab_set()
        dialog.configure(fg_color=c['main_bg'])

        # 摘要
        summary_card = ctk.CTkFrame(dialog, fg_color=c['card_bg'], corner_radius=8, border_width=1, border_color=c['card_border'])
        summary_card.pack(fill="x", padx=16, pady=(16, 8))

        ctk.CTkLabel(
            summary_card, text=f"模板共需 {summary['total']} 个变量",
            font=ctk.CTkFont(size=16, weight="bold"), text_color=c['text_primary'],
        ).pack(padx=16, pady=(12, 4))

        info_frame = ctk.CTkFrame(summary_card, fg_color="transparent")
        info_frame.pack(fill="x", padx=16, pady=(4, 12))

        ctk.CTkLabel(info_frame, text=f"已匹配 {summary['matched_count']} 个  ·  ",
                     font=ctk.CTkFont(size=14), text_color=c['success']).pack(side="left")
        ctk.CTkLabel(info_frame, text=f"未匹配 {summary['unmatched_count']} 个",
                     font=ctk.CTkFont(size=14), text_color=c['danger']).pack(side="left")

        # 滚动区
        scroll = ctk.CTkScrollableFrame(dialog, fg_color=c['card_bg'], corner_radius=8, border_width=1, border_color=c['card_border'])
        scroll.pack(fill="both", expand=True, padx=16, pady=6)

        if matched:
            ctk.CTkLabel(scroll, text="✅ 已匹配",
                         font=ctk.CTkFont(size=13, weight="bold"),
                         text_color=c['success']).pack(anchor="w", padx=14, pady=(10, 4))
            for var_name, source in matched:
                src_tag = "[公式]" if source == "formula" else "[Excel]"
                ctk.CTkLabel(scroll, text=f"  {var_name}  {src_tag}",
                             font=ctk.CTkFont(size=14),
                             text_color=c['success']).pack(anchor="w", padx=24, pady=1)

        if unmatched:
            ctk.CTkLabel(scroll, text="❌ 未匹配（需手动补全或添加公式）",
                         font=ctk.CTkFont(size=13, weight="bold"),
                         text_color=c['danger']).pack(anchor="w", padx=14, pady=(16, 4))
            for var_name in unmatched:
                ctk.CTkLabel(scroll, text=f"  {var_name}",
                             font=ctk.CTkFont(size=14),
                             text_color=c['danger']).pack(anchor="w", padx=24, pady=1)

        # 底部按钮区
        btn_bar = ctk.CTkFrame(dialog, fg_color="transparent")
        btn_bar.pack(fill="x", padx=16, pady=(8, 16))

        if unmatched:
            ctk.CTkButton(btn_bar, text="自动补全缺失列到 Excel", height=34,
                          fg_color=c['success'], hover_color=c['success_hover'],
                          text_color="#ffffff", corner_radius=6,
                          font=ctk.CTkFont(size=13, weight="bold"),
                          command=lambda: self._auto_fill_missing_columns(unmatched, dialog),
                          ).pack(side="left", padx=6)

        ctk.CTkButton(btn_bar, text="关闭", width=100, height=34,
                      fg_color=c['accent'], hover_color=c['accent_hover'],
                      corner_radius=6,
                      font=ctk.CTkFont(size=13, weight="bold"),
                      command=dialog.destroy).pack(side="right", padx=6)

    def _auto_fill_missing_columns(self, unmatched, dialog):
        """自动将缺失列以空值补入 Excel 并重新保存"""
        if self.df is None:
            return
        for col in unmatched:
            if col not in self.df.columns:
                self.df[col] = ''
        # 重新保存处理后的 Excel
        from openpyxl.workbook import Workbook
        wb = Workbook()
        ws = wb.active
        headers = list(self.df.columns)
        for ci, h in enumerate(headers, 1):
            ws.cell(row=1, column=ci).value = h
        for ri, row in enumerate(self.df.itertuples(index=False), 2):
            for ci, val in enumerate(row, 1):
                cell = ws.cell(row=ri, column=ci)
                cell.value = str(val) if pd.notna(val) else ''
                cell.number_format = '@'
        save_path = self.processed_excel_path
        wb.save(save_path)
        dialog.destroy()
        # 刷新预览
        self.load_preview()
        messagebox.showinfo("补全完成",
            f"已将 {len(unmatched)} 个缺失列补入 Excel：\n{save_path}\n\n请重新填写这些列的数据后再次处理。")

    def _run_variable_validation(self):
        """处理后运行变量验证（所有模板通用）"""
        if not self.placeholders or self.df is None:
            return
        from models.formula_engine import match_variables
        excel_cols = list(self.df.columns)
        # 所有 Excel 列都是已匹配的（包括计算生成的列）
        matched = []
        unmatched = []
        for ph in self.placeholders:
            if ph in excel_cols:
                matched.append((ph, "excel"))
            else:
                unmatched.append(ph)
        summary = {
            'total': len(self.placeholders),
            'matched_count': len(matched),
            'unmatched_count': len(unmatched),
        }
        self.show_variable_validation(matched, unmatched, summary)

    # ====================== 身份证处理弹窗（通用/自定义模板） ======================

    def _show_id_processing_dialog(self):
        """身份证处理配置弹窗（仅通用填充型/自定义公式型）。

        Returns: (dlg, result_list) — dlg 用于 wait_window()，result_list[0] 为 id_config 或 None。
        """
        c = self.colors
        dlg = ctk.CTkToplevel(self)
        dlg.title("身份证号处理")
        dlg.geometry("500x420")
        dlg.resizable(False, False)
        dlg.grab_set()
        dlg.configure(fg_color=c['main_bg'])

        card = ctk.CTkFrame(dlg, fg_color=c['card_bg'], corner_radius=10,
                            border_width=1, border_color=c['card_border'])
        card.pack(fill="both", expand=True, padx=16, pady=16)

        ctk.CTkLabel(card, text="身份证号处理",
                     font=ctk.CTkFont(family="Microsoft YaHei", size=18, weight="bold"),
                     text_color=c['accent']).pack(pady=(16, 4))
        ctk.CTkLabel(card, text="自动从身份证号提取性别、出生年月日",
                     font=ctk.CTkFont(size=13), text_color=c['text_secondary']).pack(pady=(0, 10))

        ctk.CTkFrame(card, height=1, fg_color=c['card_border']).pack(fill="x", padx=20)

        # === 控件区 ===
        ctrl = ctk.CTkFrame(card, fg_color="transparent")
        ctrl.pack(fill="both", expand=True, padx=16, pady=12)

        # 勾选框 1：是否处理
        cb1_var = ctk.BooleanVar(value=True)
        cb1 = ctk.CTkCheckBox(ctrl, text="是否处理身份证号？（自动提取性别、出生年月日）",
                              variable=cb1_var, font=ctk.CTkFont(size=14),
                              text_color=c['text_primary'])
        cb1.pack(anchor="w", pady=(6, 8))

        # 勾选框 2：是否拆分（只有一个单元格有多个身份证号）
        cb2_var = ctk.BooleanVar(value=True)
        cb2 = ctk.CTkCheckBox(ctrl, text="是否有多个身份证号存放在同一个单元格？\n（用顿号分隔，如：3303...、3303...）",
                              variable=cb2_var, font=ctk.CTkFont(size=14),
                              text_color=c['text_primary'])
        cb2.pack(anchor="w", pady=(0, 8))

        # 源列名区域
        src_frame = ctk.CTkFrame(ctrl, fg_color="transparent")
        src_frame.pack(fill="x", pady=(4, 6))

        ctk.CTkLabel(src_frame, text="源列名：",
                     font=ctk.CTkFont(size=14), text_color=c['text_primary']).pack(side="left", padx=(0, 6))

        # 自动检测 "身份证号" 列是否存在
        default_col = ""
        # 自动检测 身份证号一/二/三/四 列是否存在
        auto_id_cols = []
        if self.original_excel_path and os.path.exists(self.original_excel_path):
            try:
                ec = list(pd.read_excel(self.original_excel_path).columns)
                if "身份证号" in ec:
                    default_col = "身份证号"
                for suffix in ['一', '二', '三', '四']:
                    col = f'身份证号{suffix}'
                    if col in ec:
                        auto_id_cols.append(col)
            except Exception:
                pass

        src_var = ctk.StringVar(value=default_col)
        src_entry = ctk.CTkEntry(src_frame, textvariable=src_var, width=180,
                                 font=ctk.CTkFont(size=14),
                                 fg_color=c['entry_bg'], border_color=c['entry_border'])
        src_entry.pack(side="left")

        ctk.CTkLabel(src_frame, text="（找不到请手动输入）",
                     font=ctk.CTkFont(size=12), text_color=c['text_secondary']).pack(side="left", padx=6)

        # 手动指定列名区域（仅当 cb2 不勾选时显示）
        manual_frame = ctk.CTkFrame(ctrl, fg_color="transparent")
        manual_label = ctk.CTkLabel(manual_frame,
            text="请手动输入所有身份证号列名（一行一个）：",
            font=ctk.CTkFont(size=14), text_color=c['text_primary'])
        manual_label.pack(anchor="w", pady=(0, 4))

        manual_text = ctk.CTkTextbox(manual_frame, height=80,
                                     font=ctk.CTkFont(family="Consolas", size=14),
                                     fg_color=c['entry_bg'], border_color=c['entry_border'],
                                     border_width=1, corner_radius=6, wrap="word")
        manual_text.pack(fill="x")

        # 说明文字
        hint_label = ctk.CTkLabel(ctrl, text="",
                                  font=ctk.CTkFont(size=12), text_color=c['text_secondary'],
                                  wraplength=440, justify="left")

        def update_ui(*_):
            """根据勾选状态动态更新控件可见性"""
            enabled = cb1_var.get()
            split_mode = cb2_var.get()

            cb2.configure(state="normal" if enabled else "disabled")
            src_entry.configure(state="normal" if (enabled and split_mode) else "disabled")

            if enabled and split_mode:
                # 拆分模式：显示源列名输入
                src_frame.pack(fill="x", pady=(4, 6))
                manual_frame.pack_forget()
                hint_label.pack(anchor="w", pady=(4, 0))
                if default_col:
                    hint_label.configure(
                        text=f"✅ 已自动找到「{default_col}」列，将按顿号拆分为身份证号一、身份证号二……\n"
                             f"   并自动提取性别、出生年、出生月、出生日")
                else:
                    hint_label.configure(
                        text="⚠️ 未找到默认的「身份证号」列，请手动输入列名后点击确认")
            elif enabled and not split_mode:
                # 非拆分模式：优先自动检测已有身份证号一/二/三/四列
                src_frame.pack_forget()
                if auto_id_cols:
                    # 已自动找到：隐藏手动输入区，直接显示结果
                    manual_frame.pack_forget()
                    hint_label.pack(anchor="w", pady=(4, 0))
                    hint_label.configure(
                        text=f"✅ 已自动找到：{', '.join(auto_id_cols)}\n"
                             f"   将直接处理以上 {len(auto_id_cols)} 列，无需手动输入")
                else:
                    # 未找到：显示手动输入区
                    manual_frame.pack(fill="x", pady=(4, 6))
                    hint_label.pack(anchor="w", pady=(4, 0))
                    hint_label.configure(
                        text="⚠️ 未找到 身份证号一/二/三/四 列，请手动输入列名（每行一个）：\n"
                             "   如：身份证号一、身份证号二（列名需与 Excel 一致）")
            else:
                # 不处理身份证号
                src_frame.pack_forget()
                manual_frame.pack_forget()
                hint_label.pack_forget()

        cb1_var.trace_add("write", update_ui)
        cb2_var.trace_add("write", update_ui)
        update_ui()  # 初始状态

        # === 确认/取消 ===
        result = [None]

        def on_confirm():
            enabled = cb1_var.get()
            split_mode = cb2_var.get()

            if not enabled:
                result[0] = {'enabled': False, 'split_enabled': False,
                             'source_column': '', 'manual_columns': []}
                dlg.destroy()
                return

            if split_mode:
                col = src_var.get().strip()
                if not col:
                    messagebox.showwarning("提示", "请输入身份证号所在的列名", parent=dlg)
                    return
                # 验证列名在 Excel 中存在
                try:
                    ec = list(pd.read_excel(self.original_excel_path).columns)
                except Exception:
                    ec = []
                if col not in ec:
                    messagebox.showwarning("列名不存在",
                        f"Excel 中未找到「{col}」列，请检查列名后重试。\n\n"
                        f"Excel 现有列：{', '.join(ec[:10])}{'...' if len(ec) > 10 else ''}",
                        parent=dlg)
                    return
                result[0] = {'enabled': True, 'split_enabled': True,
                             'source_column': col, 'manual_columns': []}
            else:
                # 非拆分模式：优先使用自动检测到的列
                if auto_id_cols:
                    result[0] = {'enabled': True, 'split_enabled': False,
                                 'source_column': '', 'manual_columns': auto_id_cols}
                else:
                    text = manual_text.get("1.0", "end-1c").strip()
                    cols = [c.strip() for c in text.split('\n') if c.strip()]
                    if not cols:
                        messagebox.showwarning("提示", "请至少输入一个身份证号列名", parent=dlg)
                        return
                    # 验证列名存在
                    try:
                        ec = list(pd.read_excel(self.original_excel_path).columns)
                    except Exception:
                        ec = []
                    missing = [c for c in cols if c not in ec]
                    if missing:
                        messagebox.showwarning("列名不存在",
                            f"以下列在 Excel 中未找到：{', '.join(missing)}\n\n"
                            f"请检查列名后重试。\nExcel 现有列：{', '.join(ec[:10])}{'...' if len(ec) > 10 else ''}",
                            parent=dlg)
                        return
                    result[0] = {'enabled': True, 'split_enabled': False,
                                 'source_column': '', 'manual_columns': cols}

            dlg.destroy()

        def on_cancel():
            dlg.destroy()

        btn_frame = ctk.CTkFrame(card, fg_color="transparent")
        btn_frame.pack(fill="x", padx=16, pady=(0, 14))
        ctk.CTkButton(btn_frame, text="确认", height=36, fg_color=c['accent'],
                      hover_color=c['accent_hover'], corner_radius=6,
                      font=ctk.CTkFont(size=15, weight="bold"),
                      command=on_confirm).pack(side="left", padx=6)
        ctk.CTkButton(btn_frame, text="取消", height=36, fg_color=c['sidebar_hover'],
                      hover_color=c['accent'], text_color="#ffffff", corner_radius=6,
                      font=ctk.CTkFont(size=15, weight="bold"),
                      command=on_cancel).pack(side="right", padx=6)

        return dlg, result

    def open_folder(self, path):
        if platform.system() == "Windows":
            os.startfile(path)
        elif platform.system() == "Darwin":
            subprocess.run(["open", path])
        else:
            subprocess.run(["xdg-open", path])

if __name__ == "__main__":
    app = PaymentApp()
    app.mainloop()