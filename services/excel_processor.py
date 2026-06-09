"""
Excel 处理服务 — 读取、解析、计算、保存。

将 pandas 数据处理逻辑与 GUI 解耦。
所有函数接收回调来报告进度和检查取消状态。
"""

import pandas as pd
from openpyxl.workbook import Workbook
import config
from models.id_parser import parse_id_card
from models.fee_calculator import get_calculator
from models.formula_engine import process_formulas


def _safe_id_str(val):
    """将身份证号/电话号安全转为纯数字字符串，处理科学记数法。"""
    if val is None or (isinstance(val, float) and pd.isna(val)):
        return ''
    if isinstance(val, (int, float)):
        # 处理科学记数法：如 3.303041986072324e+17 → 完整18位字符串
        s = f'{val:.0f}'
        return s
    return str(val).strip()


def process_raw_excel(input_path, output_path, template_type,
                      formulas, placeholders,
                      progress_callback=None, cancel_checker=None,
                      id_config=None):
    """处理原始 Excel 并生成处理后的 Excel。

    Args:
        input_path: 原始 Excel 文件路径
        output_path: 处理后 Excel 保存路径
        template_type: 模板类型字符串
        custom_params: 自定义参数字典
        placeholders: Word 模板占位符列表
        progress_callback: callable(progress_float, text_str) — 进度回调
        cancel_checker: callable() -> bool — 返回 True 表示用户已取消

    Returns:
        处理后的 pandas DataFrame

    Raises:
        InterruptedError: 用户取消操作
    """
    df = pd.read_excel(input_path)
    original_columns = set(df.columns)

    # === 身份证/电话列强制文本格式（防止科学记数法） ===
    for col in df.columns:
        col_str = str(col)
        if '身份证' in col_str or '电话' in col_str:
            # 先将所有值转为字符串
            df[col] = df[col].apply(lambda x: _safe_id_str(x))

    # === 身份证列科学记数法二次检测 ===
    id_cols = [c for c in df.columns if '身份证' in str(c)]
    bad_ids = []
    for col in id_cols:
        for idx, val in df[col].items():
            s = str(val).strip().upper()
            if 'E+' in s and len(s) < 18:
                bad_ids.append((idx + 1, col, s))
    if bad_ids:
        import warnings
        detail = "\n".join([f"  第{r}行 {c}: {v}" for r, c, v in bad_ids[:5]])
        if len(bad_ids) > 5:
            detail += f"\n  ... 共 {len(bad_ids)} 处"
        raise ValueError(
            f"检测到身份证列存在科学记数法格式，将导致解析错误！\n\n"
            f"{detail}\n\n"
            f"请先打开 Excel，将身份证列设为「文本」格式后重新保存。\n"
            f"操作：选中身份证列 → 右键 → 设置单元格格式 → 文本 → 保存"
        )

    # 初始化目标列
    target_cols = [
        '被申请人姓名一', '身份证地址一', '性别一', '出生年一', '出生月一', '出生日一',
        '被申请人姓名二', '身份证地址二', '性别二', '出生年二', '出生月二', '出生日二',
    ]
    for col in target_cols:
        if col not in df.columns:
            df[col] = ''

    total_rows = len(df)

    for idx, row in df.iterrows():
        # 检查取消
        if cancel_checker and cancel_checker():
            raise InterruptedError("用户取消操作")

        # 姓名拆分
        name_str = str(row.get('业主姓名', '')).strip()
        names = [n.strip() for n in name_str.split('、') if n.strip()]
        df.at[idx, '被申请人姓名一'] = names[0] if len(names) >= 1 else ''
        df.at[idx, '被申请人姓名二'] = names[1] if len(names) >= 2 else ''

        # 身份证地址拆分
        if '身份证地址一' not in original_columns or '身份证地址二' not in original_columns:
            addr_str = str(row.get('身份证地址', '')).strip()
            addrs = [a.strip() for a in addr_str.split('、') if a.strip()]
            df.at[idx, '身份证地址一'] = addrs[0] if len(addrs) >= 1 else ''
            if len(addrs) >= 2:
                # 智能识别"同地址"、"地址同上"、"地址同前"等表述，自动复用地址一
                addr2 = addrs[1]
                same_addr_keywords = ['同地址', '地址同上', '地址同前', '同上', '同前', '同上地址', '同一地址']
                if any(kw in addr2 for kw in same_addr_keywords):
                    df.at[idx, '身份证地址二'] = addrs[0]
                else:
                    df.at[idx, '身份证地址二'] = addr2
            else:
                df.at[idx, '身份证地址二'] = addrs[0] if addrs else ''

        # 身份证解析
        id1 = str(row.get('身份证号一', '')).strip()
        if len(id1) == 18:
            y, m, d, g = parse_id_card(id1)
            df.at[idx, '出生年一'] = y
            df.at[idx, '出生月一'] = m
            df.at[idx, '出生日一'] = d
            df.at[idx, '性别一'] = g

        id2 = str(row.get('身份证号二', '')).strip()
        if len(id2) == 18:
            y, m, d, g = parse_id_card(id2)
            df.at[idx, '出生年二'] = y
            df.at[idx, '出生月二'] = m
            df.at[idx, '出生日二'] = d
            df.at[idx, '性别二'] = g

        if progress_callback:
            current_progress = (idx + 1) / total_rows
            progress_callback(current_progress, f"{int(current_progress * 100)}%")

    df = df.fillna('')

    # === 身份证号取值（id_config 驱动） ===
    if id_config and id_config.get('enabled'):
        if id_config.get('split_enabled'):
            # 顿号拆分模式：将源列按 "、" 拆分为多个身份证号列
            source_col = id_config.get('source_column', '')
            if source_col and source_col in df.columns:
                # 确保身份证号一/二 列存在
                for n in range(1, 5):
                    suffix = '一' if n == 1 else ('二' if n == 2 else ('三' if n == 3 else '四'))
                    col_name = f'身份证号{suffix}'
                    if col_name not in df.columns:
                        df[col_name] = ''
                for idx, row in df.iterrows():
                    raw = str(row.get(source_col, '')).strip()
                    if '、' in raw:
                        parts = [p.strip() for p in raw.split('、') if p.strip()]
                        for n, part in enumerate(parts[:4], 1):
                            suffix = '一' if n == 1 else ('二' if n == 2 else ('三' if n == 3 else '四'))
                            df.at[idx, f'身份证号{suffix}'] = part
                    else:
                        # 单个身份证号也填入身份证号一
                        if raw and len(raw) == 18:
                            df.at[idx, '身份证号一'] = raw
        else:
            # 手动指定列模式：列名已在 manual_columns 中，确保存在
            pass

        # 统一解析所有 身份证号一/二/三/四 列 → 性别/出生年月日
        for n in range(1, 5):
            suffix = '一' if n == 1 else ('二' if n == 2 else ('三' if n == 3 else '四'))
            id_col = f'身份证号{suffix}'
            gender_col = f'性别{suffix}'
            year_col = f'出生年{suffix}'
            month_col = f'出生月{suffix}'
            day_col = f'出生日{suffix}'
            for col in [gender_col, year_col, month_col, day_col]:
                if col not in df.columns:
                    df[col] = ''
            if id_col in df.columns:
                for idx, row in df.iterrows():
                    id_val = str(row.get(id_col, '')).strip()
                    if len(id_val) == 18:
                        try:
                            y, m, d, g = parse_id_card(id_val)
                            df.at[idx, gender_col] = g
                            df.at[idx, year_col] = y
                            df.at[idx, month_col] = m
                            df.at[idx, day_col] = d
                        except Exception:
                            pass

    # 格式化电话号码（去除 .0 后缀）
    for col in ['联系电话一', '联系电话二']:
        if col in df.columns:
            df[col] = df[col].astype(str).str.replace(r'\.0$', '', regex=True)

    # 智能单价处理（仅瓯海和平阳模板）
    if template_type in ["预设公式型（瓯海）", "预设公式型（平阳）"]:
        price_col_exists = '单价' in df.columns
        price_col_has_data = price_col_exists and df['单价'].astype(str).str.strip().ne('').any()
        if not price_col_has_data:
            if '房屋类型' in df.columns:
                if not price_col_exists:
                    df['单价'] = ''
                for idx, row in df.iterrows():
                    house_type = str(row.get('房屋类型', '')).strip()
                    default_price = config.HOUSE_TYPE_PRICE_MAP.get(house_type)
                    if default_price:
                        df.at[idx, '单价'] = default_price

    # 费用计算
    if template_type == "自定义公式型":
        # 公式引擎模式
        if formulas:
            df, _, _ = process_formulas(df, formulas)
    else:
        # 策略模式（瓯海/平阳/通用）
        calculator = get_calculator(template_type)
        calculator.validate_columns(df)

        if template_type != "通用填充型":
            calc_results = df.apply(calculator.calculate, axis=1, result_type='expand')
            for col in calc_results.columns:
                df[col] = calc_results[col]

    # 补全占位符列
    for ph in placeholders:
        if ph not in df.columns:
            df[ph] = ''

    # 保存 Excel（强制文本格式）
    wb = Workbook()
    ws = wb.active
    headers = list(df.columns)
    for col_idx, header in enumerate(headers, 1):
        ws.cell(row=1, column=col_idx).value = header
    for r_idx, row in enumerate(df.itertuples(index=False), 2):
        for c_idx, value in enumerate(row, 1):
            cell = ws.cell(row=r_idx, column=c_idx)
            cell.value = str(value) if pd.notna(value) else ''
            cell.number_format = '@'
    wb.save(output_path)

    return df
